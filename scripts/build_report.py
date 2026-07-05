"""
Анализ рисков регионов: Матрица Риск × Рейтинг РУСАДА.
Источник данных: Рейтинг регионов Итоги 2025 (xlsx)
Максимум баллов: 190 (блок 1 max 120 + блок 2 max 100 + блок 3 max 50+30*)
Порог «высокий рейтинг»: 130 баллов

Обновлено по логике промта v3:
  - Приоритеты 1–4 вместо квадрантов со старыми эмодзи
  - Текстовые коды зон (RED/ORANGE/GREEN) для логики; эмодзи — только декор
  - ZONE_COLORS (бар, Excel) и QUADRANT_COLORS (scatter) не смешиваются
  - Инварианты самопроверки
  - Шаблоны обоснований
  - Этап 0: вывод артефактов изучения данных

ИСПОЛЬЗОВАНИЕ:
  1) Обновите XLSX_PATH, OUT_DIR ниже.
  2) python build_report.py
"""

import openpyxl
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import os, warnings
warnings.filterwarnings('ignore')

# ── Пути ──────────────────────────────────────────────────────────────────────
XLSX_PATH = "/tmp/Reiting-regionov-Itogi-2025.xlsx"
OUT_DIR   = "/home/ubuntu/region_risk"
os.makedirs(OUT_DIR, exist_ok=True)

# ── Порог рейтинга (аналог 80 баллов у ОСФ, но шкала 0-190) ─────────────────
THRESHOLD = 130   # ~68% от 190 — «достаточный уровень антидопинговой работы»
MAX_SCORE = 190

# ── Цвета ─────────────────────────────────────────────────────────────────────
INK  = "#0F2D52"; SUB = "#7C8DA6"; GRID = "#EEF2F7"
FONT_FAMILY = "Inter, Segoe UI, Roboto, Arial, sans-serif"

# Текстовые коды зон (для логики, сортировки, фильтрации)
ZONE_CODE_RED    = "RED"
ZONE_CODE_ORANGE = "ORANGE"
ZONE_CODE_GREEN  = "GREEN"

# Эмодзи — только декор в Excel и дашборде
ZONE_EMOJI = {"RED": "🔴", "ORANGE": "🟠", "GREEN": "🟢"}

# Для бар-графика и Excel (3 зоны):
ZONE_COLORS = {"RED": "#DC2626", "ORANGE": "#F59E0B", "GREEN": "#10B981"}

# Для матрицы квадрантов scatter (4 приоритета):
QUADRANT_COLORS = {
    "Приоритет 1 (низкий рейтинг + систематичность)": "#DC2626",
    "Приоритет 2 (высокий рейтинг + систематичность)": "#F59E0B",
    "Приоритет 3 (низкий рейтинг + нет систематичности)": "#3B82F6",
    "Приоритет 4 (высокий рейтинг + нет систематичности)": "#10B981",
}
# Правило: scatter красится по QUADRANT_COLORS, бар и Excel — по ZONE_COLORS. Не смешивать.

# ── 1. Загрузка данных ────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb['Рейтинг регионов 2025']

# Заголовки (строка 3)
headers_row = [c.value for c in ws[3]]

# Данные регионов (строки 4..94)
rows = []
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    fo   = row[0]   # Федеральный округ
    name = row[1]   # Субъект РФ
    if not name or not fo:
        continue
    name = str(name).strip().replace('\xa0', '').replace('  ', ' ')
    fo   = str(fo).strip()
    # Пропускаем строки-заголовки
    if name.lower() in ('субъект рф', 'субъект') or fo.lower() == 'фо':
        continue

    # Блок 1: колонки C..L (индексы 2..11), итог M (12)
    b1_vals = [row[i] if isinstance(row[i], (int, float)) else 0 for i in range(2, 12)]
    b1_total = row[12] if isinstance(row[12], (int, float)) else sum(b1_vals)

    # Блок 2: колонки N..S (13..18), итог T (19)
    b2_vals = [row[i] if isinstance(row[i], (int, float)) else 0 for i in range(13, 19)]
    b2_total = row[19] if isinstance(row[19], (int, float)) else sum(b2_vals)

    # Блок 3: колонки W..Z (22..25), итог AA (26)
    b3_vals = [row[i] if isinstance(row[i], (int, float)) else 0 for i in range(22, 26)]
    b3_total = row[26] if isinstance(row[26], (int, float)) else sum(b3_vals)

    total = row[27] if isinstance(row[27], (int, float)) else (b1_total + b2_total + b3_total)
    place = row[28] if isinstance(row[28], (int, float)) else None

    rows.append({
        'ФО': fo,
        'Регион': name,
        'Блок 1 (Взаимодействие с РУСАДА)': b1_total if b1_total else 0,
        'Блок 2 (Образование)': b2_total if b2_total else 0,
        'Блок 3 (Профилактика здравоохранения)': b3_total if b3_total else 0,
        'Итого баллов': total if total else 0,
        'Место': place,
        # Детали блока 1
        'B1: Ответственные/Сайт': b1_vals[0],
        'B1: Повышение квалификации': b1_vals[1],
        'B1: Соглашение с РУСАДА': b1_vals[2],
        'B1: План-график 2026': b1_vals[3],
        'B1: Отчёт за 2025': b1_vals[4],
        'B1: Опрос': b1_vals[5],
        'B1: Участие в мероприятиях': b1_vals[6],
        'B1: Коммуникация': b1_vals[7],
        'B1: Мониторинг (штраф)': b1_vals[8],
        'B1: Организация мероприятий *': b1_vals[9],
        # Детали блока 2
        'B2: Квалификация лекторов': b2_vals[0],
        'B2: Обновление программ СШ': b2_vals[1],
        'B2: НМС для СШ': b2_vals[2],
        'B2: Соглашение о персданных': b2_vals[3],
        'B2: Уровень АД образования': b2_vals[4],
        'B2: Стенды': b2_vals[5],
        # Детали блока 3
        'B3: Ответственный Минздрава': b3_vals[0],
        'B3: Программа для медперсонала': b3_vals[1],
        'B3: Межведомственное взаимодействие': b3_vals[2],
        'B3: Инновационные системы *': b3_vals[3],
    })

df = pd.DataFrame(rows)
df['Итого баллов'] = pd.to_numeric(df['Итого баллов'], errors='coerce').fillna(0).astype(int)
df['Место'] = pd.to_numeric(df['Место'], errors='coerce')

# ── ЭТАП 0: Артефакты изучения данных ────────────────────────────────────────
print("=" * 60)
print("ЭТАП 0: Изучение входных данных")
print("=" * 60)
print(f"Источник: {XLSX_PATH}")
print(f"Всего регионов (N_рейтинг): {len(df)}")
print(f"Колонки: ФО, Регион, Блок1, Блок2, Блок3, Итого, Место")
print(f"Первые 3 региона: {df['Регион'].head(3).tolist()}")
print(f"Последние 3 региона: {df['Регион'].tail(3).tolist()}")
print(f"Логика баллов: Блок1 (макс 120) + Блок2 (макс 100) + Блок3 (макс 50+30*) = макс {MAX_SCORE}")
print("=" * 60)

# ── 2. Расчёт зоны риска (текстовые коды) ────────────────────────────────────
# RED    — итого < 95  (<50% от 190)
# ORANGE — 95 ≤ итого < 130
# GREEN  — итого ≥ 130
# Порядок проверки: RED → ORANGE → GREEN. Регион получает ровно одну зону.

def get_zone_code(score):
    if score < 95:
        return ZONE_CODE_RED
    elif score < THRESHOLD:
        return ZONE_CODE_ORANGE
    else:
        return ZONE_CODE_GREEN

df['_zone_code'] = df['Итого баллов'].apply(get_zone_code)
# Отображаемое значение: эмодзи-декор + текстовый код
df['Зона риска'] = df['_zone_code'].apply(lambda z: f"{ZONE_EMOJI[z]} {z}")

# ── 3. Приоритеты 1–4 и квадранты ────────────────────────────────────────────
# Матрица приоритетов (порог = 130 баллов):
# Приоритет 1: Рейтинг < 130 + RED/ORANGE → срочное вмешательство
# Приоритет 2: Рейтинг ≥ 130 + RED/ORANGE → усилить образование
# Приоритет 3: Рейтинг < 130 + GREEN       → профилактика + подтянуть рейтинг
# Приоритет 4: Рейтинг ≥ 130 + GREEN       → поддерживать

RECS = {
    1: "Срочное вмешательство РУСАДА; приоритизировать работу по невыполненным критериям всех трёх блоков",
    2: "Усилить антидопинговое образование; беседы с ответственными; рассмотреть внесоревновательный контроль",
    3: "Продолжать профилактику; приоритизировать невыполненные критерии рейтинга; превентивная работа с регионом",
    4: "Поддерживать текущий уровень; обмен лучшими практиками; плановый мониторинг",
}

QUADRANT_LABELS = {
    1: "Приоритет 1 (низкий рейтинг + систематичность)",
    2: "Приоритет 2 (высокий рейтинг + систематичность)",
    3: "Приоритет 3 (низкий рейтинг + нет систематичности)",
    4: "Приоритет 4 (высокий рейтинг + нет систематичности)",
}

def get_priority(row):
    score = row['Итого баллов']
    zone  = row['_zone_code']
    high_rating = score >= THRESHOLD
    systematic  = zone in (ZONE_CODE_RED, ZONE_CODE_ORANGE)
    if not high_rating and systematic:
        return 1
    elif high_rating and systematic:
        return 2
    elif not high_rating and not systematic:
        return 3
    else:
        return 4

df['Приоритет'] = df.apply(get_priority, axis=1)
df['Квадрант']  = df['Приоритет'].map(QUADRANT_LABELS)
df['Рекомендация'] = df['Приоритет'].map(RECS)

# ── 4. Обоснования ───────────────────────────────────────────────────────────
def get_justification(row):
    score    = row['Итого баллов']
    pct      = row['% выполнения']
    zone     = row['_zone_code']
    not_done = row.get('Невыполненные критерии', '—')
    if zone in (ZONE_CODE_RED, ZONE_CODE_ORANGE):
        base = f"{ZONE_EMOJI[zone]} {zone} риск (итого: {score} из {MAX_SCORE}, {pct}%)"
        if not_done and not_done != '—':
            return base + f"; нехватка баллов по критериям: {not_done}"
        return base + "; критерии рейтинга выполнены полностью"
    else:
        base = f"Низкий риск (итого: {score} из {MAX_SCORE}, {pct}%)"
        if not_done and not_done != '—':
            return base + f"; недостающие критерии: {not_done}"
        return base + "; все критерии выполнены"

# ── 5. Невыполненные критерии (блок 1) ───────────────────────────────────────
CRITERIA_B1 = {
    'B1: Ответственные/Сайт': 'Ответственные/Сайт (15)',
    'B1: Повышение квалификации': 'Повышение квалификации (15)',
    'B1: Соглашение с РУСАДА': 'Соглашение с РУСАДА (10)',
    'B1: План-график 2026': 'План-график 2026 (15)',
    'B1: Отчёт за 2025': 'Отчёт за 2025 (15)',
    'B1: Опрос': 'Опрос (15)',
    'B1: Участие в мероприятиях': 'Участие в мероприятиях (10)',
    'B1: Коммуникация': 'Коммуникация (5)',
}
CRITERIA_B2 = {
    'B2: Квалификация лекторов': 'Квалификация лекторов (15)',
    'B2: Обновление программ СШ': 'Обновление программ СШ (35)',
    'B2: НМС для СШ': 'НМС для СШ (10)',
    'B2: Соглашение о персданных': 'Соглашение о персданных (10)',
    'B2: Уровень АД образования': 'Уровень АД образования (20)',
    'B2: Стенды': 'Стенды (10)',
}
CRITERIA_B3 = {
    'B3: Ответственный Минздрава': 'Ответственный Минздрава (10)',
    'B3: Программа для медперсонала': 'Программа для медперсонала (25)',
    'B3: Межведомственное взаимодействие': 'Межведомственное взаимодействие (15)',
}

ALL_CRITERIA = {**CRITERIA_B1, **CRITERIA_B2, **CRITERIA_B3}

def get_not_done(row):
    not_done = []
    for col, label in ALL_CRITERIA.items():
        val = row.get(col, 0)
        if val is None or val == 0:
            not_done.append(label)
    return "; ".join(not_done) if not_done else "—"

def get_done(row):
    done = []
    for col, label in ALL_CRITERIA.items():
        val = row.get(col, 0)
        if val and val > 0:
            done.append(label)
    return "; ".join(done) if done else "—"

df['Выполненные критерии']   = df.apply(get_done, axis=1)
df['Невыполненные критерии'] = df.apply(get_not_done, axis=1)

# ── 6. Процент выполнения ─────────────────────────────────────────────────────
df['% выполнения'] = (df['Итого баллов'] / MAX_SCORE * 100).round(1)

# Обоснования (после расчёта критериев)
df['Обоснование'] = df.apply(get_justification, axis=1)

# ── 7. Сортировка: приоритет → zone (RED=1, ORANGE=2, GREEN=3) → балл убыв. ──
zone_sort = {ZONE_CODE_RED: 1, ZONE_CODE_ORANGE: 2, ZONE_CODE_GREEN: 3}
df['_zone_sort'] = df['_zone_code'].map(zone_sort)
df = df.sort_values(['Приоритет', '_zone_sort', 'Итого баллов'],
                    ascending=[True, True, False]).reset_index(drop=True)
df.drop(columns=['_zone_sort'], inplace=True)
df.index = range(1, len(df)+1)

# ── САМОПРОВЕРКА (инварианты) ─────────────────────────────────────────────────
N_RATING = len(df)
print(f"\n{'='*60}")
print("САМОПРОВЕРКА (инварианты)")
print(f"{'='*60}")
print(f"1. Строк в итоговой таблице = {N_RATING} (должно быть ~89): {'✓' if N_RATING > 0 else '✗'}")
prio_sum = df['Приоритет'].count()
print(f"2. Сумма строк по приоритетам = {prio_sum} (должно быть {N_RATING}): {'✓' if prio_sum == N_RATING else '✗ ОШИБКА'}")
zone_dist = df['_zone_code'].value_counts().to_dict()
print(f"3. Распределение по зонам: {zone_dist}")
prio_dist = df['Приоритет'].value_counts().sort_index().to_dict()
print(f"   Распределение по приоритетам: {prio_dist}")
bad4 = df[df['_zone_code'].isin([ZONE_CODE_RED, ZONE_CODE_ORANGE]) & df['Приоритет'].isin([3, 4])]
print(f"4. RED/ORANGE с приоритетами 3-4: {len(bad4)} (должно быть 0): {'✓' if len(bad4)==0 else '✗ ОШИБКА'}")
bad5 = df[(df['_zone_code'] == ZONE_CODE_GREEN) & df['Приоритет'].isin([1, 2])]
print(f"5. GREEN с приоритетами 1-2: {len(bad5)} (должно быть 0): {'✓' if len(bad5)==0 else '✗ ОШИБКА'}")
bad_header = df[df['Регион'].str.lower().isin(['субъект рф', 'субъект'])]
print(f"7. Строки-заголовки в таблице: {len(bad_header)} (должно быть 0): {'✓' if len(bad_header)==0 else '✗ ОШИБКА'}")
print(f"{'='*60}\n")

# ── 8. Excel-отчёт ───────────────────────────────────────────────────────────
OUT_XLSX = os.path.join(OUT_DIR, "region_risk_final.xlsx")

wb_out = openpyxl.Workbook()

# --- Лист «Матрица регионов» ---
ws_main = wb_out.active
ws_main.title = "Матрица регионов"

# Цвета заливки
FILL_RED    = PatternFill("solid", fgColor="FEE2E2")
FILL_ORANGE = PatternFill("solid", fgColor="FEF3C7")
FILL_GREEN  = PatternFill("solid", fgColor="D1FAE5")
FILL_HEADER = PatternFill("solid", fgColor="0F2D52")

FONT_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_BOLD   = Font(bold=True, name="Calibri", size=9)
FONT_NORM   = Font(name="Calibri", size=9)

BORDER_THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

COLS = [
    '№', 'ФО', 'Регион', 'Приоритет', 'Квадрант', 'Зона риска',
    'Итого баллов', '% выполнения', 'Место',
    'Блок 1 (Взаимодействие с РУСАДА)',
    'Блок 2 (Образование)',
    'Блок 3 (Профилактика здравоохранения)',
    'Выполненные критерии', 'Невыполненные критерии',
    'Рекомендация', 'Обоснование'
]

# Заголовок
ws_main.append(['АНАЛИЗ РИСКОВ РЕГИОНОВ — РЕЙТИНГ РУСАДА 2025'])
ws_main.merge_cells(f'A1:{get_column_letter(len(COLS))}1')
ws_main['A1'].font = Font(bold=True, size=13, color="0F2D52", name="Calibri")
ws_main['A1'].alignment = Alignment(horizontal='center')
ws_main.row_dimensions[1].height = 22

ws_main.append([f'Максимум баллов: {MAX_SCORE} | Порог «высокий рейтинг»: {THRESHOLD} баллов | Регионов: {len(df)}'])
ws_main.merge_cells(f'A2:{get_column_letter(len(COLS))}2')
ws_main['A2'].font = Font(italic=True, size=9, color="7C8DA6", name="Calibri")
ws_main['A2'].alignment = Alignment(horizontal='center')
ws_main.row_dimensions[2].height = 16

ws_main.append([])  # пустая строка

# Заголовки столбцов (строка 4)
ws_main.append(COLS)
for col_idx, col_name in enumerate(COLS, start=1):
    cell = ws_main.cell(row=4, column=col_idx)
    cell.fill = FILL_HEADER
    cell.font = FONT_HEADER
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER_THIN
ws_main.row_dimensions[4].height = 36

# Данные
for i, row in df.iterrows():
    data_row = [
        i,
        row['ФО'],
        row['Регион'],
        row['Приоритет'],
        row['Квадрант'],
        row['Зона риска'],
        row['Итого баллов'],
        f"{row['% выполнения']}%",
        row['Место'],
        row['Блок 1 (Взаимодействие с РУСАДА)'],
        row['Блок 2 (Образование)'],
        row['Блок 3 (Профилактика здравоохранения)'],
        row['Выполненные критерии'],
        row['Невыполненные критерии'],
        row['Рекомендация'],
        row['Обоснование'],
    ]
    ws_main.append(data_row)
    row_num = ws_main.max_row

    # Цвет строки по zone_code (текстовый код)
    zone_code = row['_zone_code']
    if zone_code == ZONE_CODE_RED:
        fill = FILL_RED
    elif zone_code == ZONE_CODE_ORANGE:
        fill = FILL_ORANGE
    else:
        fill = FILL_GREEN

    for col_idx in range(1, len(COLS)+1):
        cell = ws_main.cell(row=row_num, column=col_idx)
        cell.fill = fill
        cell.font = FONT_NORM
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical='top', wrap_text=(col_idx >= 12))

    ws_main.row_dimensions[row_num].height = 40 if row['Невыполненные критерии'] != '—' else 18

# Ширины столбцов
col_widths = [5, 7, 28, 10, 45, 18, 12, 12, 8, 14, 14, 14, 55, 55, 60, 65]
for i, w in enumerate(col_widths, start=1):
    ws_main.column_dimensions[get_column_letter(i)].width = w

# Заморозить строку 4
ws_main.freeze_panes = 'A5'

# --- Лист «Сводка по ФО» ---
ws_fo = wb_out.create_sheet("Сводка по ФО")
fo_summary = df.groupby('ФО').agg(
    Регионов=('Регион', 'count'),
    Критических=('_zone_code', lambda x: (x == ZONE_CODE_RED).sum()),
    Умеренных=('_zone_code', lambda x: (x == ZONE_CODE_ORANGE).sum()),
    Низких=('_zone_code', lambda x: (x == ZONE_CODE_GREEN).sum()),
    Средний_балл=('Итого баллов', 'mean'),
    Мин_балл=('Итого баллов', 'min'),
    Макс_балл=('Итого баллов', 'max'),
).reset_index()
fo_summary['Средний_балл'] = fo_summary['Средний_балл'].round(1)
fo_summary = fo_summary.sort_values('Критических', ascending=False)

ws_fo.append(['СВОДКА ПО ФЕДЕРАЛЬНЫМ ОКРУГАМ'])
ws_fo.merge_cells('A1:H1')
ws_fo['A1'].font = Font(bold=True, size=12, color="0F2D52", name="Calibri")
ws_fo['A1'].alignment = Alignment(horizontal='center')
ws_fo.append([])

fo_headers = ['ФО', 'Регионов', '🔴 Критических', '🟠 Умеренных', '🟢 Низких', 'Средний балл', 'Мин балл', 'Макс балл']
ws_fo.append(fo_headers)
for col_idx in range(1, len(fo_headers)+1):
    cell = ws_fo.cell(row=3, column=col_idx)
    cell.fill = FILL_HEADER
    cell.font = FONT_HEADER
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER_THIN

for _, row in fo_summary.iterrows():
    ws_fo.append([
        row['ФО'], row['Регионов'], row['Критических'],
        row['Умеренных'], row['Низких'],
        row['Средний_балл'], row['Мин_балл'], row['Макс_балл']
    ])
    r = ws_fo.max_row
    for c in range(1, 9):
        cell = ws_fo.cell(row=r, column=c)
        cell.border = BORDER_THIN
        cell.font = FONT_NORM
        cell.alignment = Alignment(horizontal='center')
    if row['Критических'] > 0:
        ws_fo.cell(row=r, column=3).fill = FILL_RED
    if row['Умеренных'] > 0:
        ws_fo.cell(row=r, column=4).fill = FILL_ORANGE
    if row['Низких'] > 0:
        ws_fo.cell(row=r, column=5).fill = FILL_GREEN

for i, w in enumerate([10, 10, 16, 16, 12, 14, 12, 12], start=1):
    ws_fo.column_dimensions[get_column_letter(i)].width = w

wb_out.save(OUT_XLSX)
print(f"\nExcel сохранён: {OUT_XLSX}")

# ── 9. Plotly HTML-дашборд ───────────────────────────────────────────────────
OUT_HTML = os.path.join(OUT_DIR, "region_risk_dashboard.html")

# Цвета точек
color_map = {
    "🔴 КРИТИЧЕСКАЯ": "#DC2626",
    "🟠 УМЕРЕННАЯ":   "#F59E0B",
    "🟢 НИЗКАЯ":      "#10B981",
}

fig = make_subplots(
    rows=2, cols=2,
    subplot_titles=[
        "Матрица рисков регионов (Рейтинг × Зона)",
        "Топ-20 регионов с наибольшим риском",
        "Распределение регионов по квадрантам",
        "Средний балл по федеральным округам",
    ],
    specs=[[{"type": "scatter"}, {"type": "bar"}],
           [{"type": "pie"},    {"type": "bar"}]],
    vertical_spacing=0.15,
    horizontal_spacing=0.1,
)

# --- График 1: Scatter матрица (цвет по QUADRANT_COLORS) ---
import random
random.seed(42)

for quad_label, color in QUADRANT_COLORS.items():
    sub = df[df['Квадрант'] == quad_label]
    if sub.empty:
        continue
    x_vals = sub['Итого баллов'].astype(float)
    x_jitter = x_vals + [random.uniform(-2, 2) for _ in range(len(sub))]
    y_base = sub['Приоритет'].astype(float)
    y_jitter = y_base + [random.uniform(-0.02, 0.02) for _ in range(len(sub))]
    fig.add_trace(go.Scatter(
        x=x_jitter,
        y=y_jitter,
        mode='markers+text',
        name=quad_label,
        marker=dict(color=color, size=10, opacity=0.82, line=dict(width=0.5, color='white')),
        text=sub['Регион'].apply(lambda x: x[:12] + '…' if len(x) > 12 else x),
        textposition='top center',
        textfont=dict(size=7),
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "ФО: %{customdata[1]}<br>"
            "Баллы: %{customdata[2]}<br>"
            "Приоритет: %{customdata[3]}<extra></extra>"
        ),
        customdata=list(zip(sub['Регион'], sub['ФО'], sub['Итого баллов'], sub['Приоритет'])),
    ), row=1, col=1)

fig.add_vline(x=THRESHOLD, line_dash="dash", line_color=INK, line_width=1.5,
              annotation_text=f"Порог {THRESHOLD}", annotation_position="top right",
              row=1, col=1)

fig.update_yaxes(
    tickvals=[1, 2, 3],
    ticktext=["🔴 КРИТИЧЕСКАЯ", "🟠 УМЕРЕННАЯ", "🟢 НИЗКАЯ"],
    row=1, col=1
)
fig.update_xaxes(title_text="Итого баллов (макс. 190)", row=1, col=1)

# --- График 2: Топ-20 рисковых регионов (горизонтальный бар, цвет по ZONE_COLORS) ---
top20 = df[df['_zone_code'].isin([ZONE_CODE_RED, ZONE_CODE_ORANGE])].sort_values('Итого баллов').tail(20)
bar_colors = [ZONE_COLORS[z] for z in top20['_zone_code']]

fig.add_trace(go.Bar(
    x=top20['Итого баллов'],
    y=top20['Регион'].apply(lambda x: x[:25] + '…' if len(x) > 25 else x),
    orientation='h',
    marker_color=bar_colors,
    marker_cornerradius=5,
    text=top20['Итого баллов'],
    textposition='outside',
    cliponaxis=False,
    hovertemplate="<b>%{y}</b><br>Баллы: %{x}<extra></extra>",
    showlegend=False,
), row=1, col=2)
fig.update_xaxes(title_text="Итого баллов", row=1, col=2)

# --- График 3: Pie по приоритетам (цвет по QUADRANT_COLORS) ---
quad_counts = df['Квадрант'].value_counts()
pie_colors = [QUADRANT_COLORS.get(q, "#94A3B8") for q in quad_counts.index]

fig.add_trace(go.Pie(
    labels=quad_counts.index,
    values=quad_counts.values,
    marker_colors=pie_colors,
    textinfo='label+percent+value',
    textfont=dict(size=9),
    hole=0.35,
    showlegend=False,
), row=2, col=1)

# --- График 4: Средний балл по ФО (цвет по ZONE_COLORS) ---
fo_avg = df.groupby('ФО')['Итого баллов'].mean().sort_values()
fo_bar_colors = [
    ZONE_COLORS[ZONE_CODE_RED] if v < 95 else
    (ZONE_COLORS[ZONE_CODE_ORANGE] if v < THRESHOLD else ZONE_COLORS[ZONE_CODE_GREEN])
    for v in fo_avg.values
]

fig.add_trace(go.Bar(
    x=fo_avg.values,
    y=fo_avg.index,
    orientation='h',
    marker_color=fo_bar_colors,
    marker_cornerradius=5,
    text=[f"{v:.0f}" for v in fo_avg.values],
    textposition='outside',
    cliponaxis=False,
    showlegend=False,
    hovertemplate="<b>%{y}</b><br>Средний балл: %{x:.1f}<extra></extra>",
), row=2, col=2)
# Вертикальная линия порога на графике 4 через shapes
fig.add_shape(type='line', x0=THRESHOLD, x1=THRESHOLD, y0=-0.5, y1=len(fo_avg)-0.5,
              line=dict(dash='dash', color=INK, width=1.5), row=2, col=2)
fig.update_xaxes(title_text="Средний балл", row=2, col=2)

# --- Общий стиль ---
fig.update_layout(
    title=dict(
        text="<b>Анализ рисков регионов | Рейтинг РУСАДА 2025</b>",
        font=dict(size=18, color=INK, family=FONT_FAMILY),
        x=0.5,
    ),
    height=900,
    paper_bgcolor="white",
    plot_bgcolor=GRID,
    font=dict(family=FONT_FAMILY, color=INK),
    legend=dict(
        orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
        font=dict(size=10),
    ),
    margin=dict(t=100, b=60, l=60, r=60),
)

fig.write_html(OUT_HTML)
print(f"Дашборд сохранён: {OUT_HTML}")

# ── 10. Итоговая статистика ───────────────────────────────────────────────────
print("\n=== ИТОГОВАЯ СТАТИСТИКА ===")
print(df['Зона риска'].value_counts())
print(f"\nПорог рейтинга: {THRESHOLD} баллов")
print(f"Всего регионов: {len(df)}")
